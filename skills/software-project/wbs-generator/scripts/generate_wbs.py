#!/usr/bin/env python3
"""
WBS Excel Generator - 根据JSON数据生成WBS Excel文件
样式严格参照电源二组WBS模板

用法:
  python generate_wbs.py <input_json> <output_xlsx>

input_json 格式见 references/wbs-data-schema.md
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("需要安装 openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


# ============================================================
# 样式常量 - 严格匹配模板
# ============================================================

# 边框
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
NO_BORDER = Border()

# 填充色 (精确匹配模板 fills)
FILL_TITLE = PatternFill(start_color="FF1450B8", end_color="FF1450B8", fill_type="solid")       # fill[2] 标题蓝
FILL_INFO = PatternFill(start_color="FFEAFAF1", end_color="FFEAFAF1", fill_type="solid")         # fill[3] 表头区浅绿
FILL_COL_HEADER = PatternFill(start_color="FF6EAB29", end_color="FF6EAB29", fill_type="solid")   # fill[4] 列标题绿
FILL_COL_HEADER_D = PatternFill(start_color="FFF5971D", end_color="FFF5971D", fill_type="solid") # fill[5] D列标题橙
FILL_SECTION = PatternFill(start_color="FFF5F5F5", end_color="FFF5F5F5", fill_type="solid")      # fill[6]/[9] 分区标题浅灰
FILL_ROW_GREEN = PatternFill(start_color="FFEAFAF1", end_color="FFEAFAF1", fill_type="solid")    # fill[7] 【硬件】行浅绿
FILL_ROW_GRAY = PatternFill(start_color="FFF3F5F7", end_color="FFF3F5F7", fill_type="solid")     # fill[8] 【工厂】行浅灰蓝

# 字体 (精确匹配模板 fonts)
# font[3] - 标题
FONT_TITLE = Font(name="Microsoft YaHei", size=24, bold=True, color="FFFFFFFF")
# font[0] - 基础等线
FONT_BASE = Font(name="等线", size=10)
# font[5] - 表头区标签加粗
FONT_INFO_LABEL = Font(name="等线", size=10, bold=True)
# font[12] - 优先等级(红色大号)
FONT_PRIORITY = Font(name="LarkHackSafariFont", size=18, bold=True, color="FFFF0000")
# font[10] - 硬件参数(蓝色)
FONT_HW_NOTES = Font(name="等线", size=10, color="FF0070C0")
# font[13] - 硬件参数标题(大号深灰)
FONT_HW_TITLE = Font(name="LarkHackSafariFont", size=16, bold=True, color="FF474747")

# font[15] - 列标题(白色加粗)
FONT_COL_HEADER = Font(name="LarkHackSafariFont", size=12, bold=True, color="FFFFFFFF")
# font[21] - 备注列标题(灰字)
FONT_COL_HEADER_E = Font(name="等线", size=12, bold=True, color="FFBFBFBF")

# font[17] - 分区标题(深绿加粗)
FONT_SECTION = Font(name="Microsoft YaHei", size=11, bold=True, color="FF447A02")
# font[18]/[19] - 分区D列提示(灰色小字)
FONT_SECTION_D = Font(name="Microsoft YaHei", size=9, color="FFC2C2C2")
# font[20]/[21] - 分区提示文字
FONT_HINT = Font(name="LarkHackSafariFont", size=9, color="FFC2C2C2")
FONT_HINT_E = Font(name="等线", size=9, color="FFBFBFBF")

# 普通任务行字体 (无底色行)
# font[22] - A列任务名(深灰)
FONT_TASK_A = Font(name="Microsoft YaHei", color="FF474747")
# E列备注(蓝色)
FONT_TASK_E = Font(name="Microsoft YaHei", color="FF0070C0")
# font[23] - B列负责人(蓝色)
FONT_TASK_B = Font(name="Microsoft YaHei", color="FF2F77F7")
# font[0]/[24] - C列状态(等线默认)
FONT_TASK_C = Font(name="等线", size=10)
# font[25] - D列日期(加粗深灰)
FONT_TASK_D = Font(name="LarkHackSafariFont", bold=True, color="FF474747")

# 【硬件】行字体 (浅绿底)
# font[36] - A列(深灰偏黑)
FONT_HW_TASK_A = Font(name="Microsoft YaHei", size=10, color="FF333333")
# font[37] - B列(蓝色)
FONT_HW_TASK_B = Font(name="Microsoft YaHei", size=10, color="FF2F77F7")
# font[0] - C列
FONT_HW_TASK_C = Font(name="等线", size=10)
# font[25] - D列
FONT_HW_TASK_D = Font(name="LarkHackSafariFont", bold=True, color="FF474747")
# E列备注(蓝色)
FONT_HW_TASK_E = Font(name="Microsoft YaHei", color="FF0070C0")

# 【工厂】行字体 (浅灰蓝底) - 同【硬件】行字体
FONT_FC_TASK_A = FONT_HW_TASK_A
FONT_FC_TASK_B = FONT_HW_TASK_B
FONT_FC_TASK_C = FONT_HW_TASK_C
FONT_FC_TASK_D = FONT_HW_TASK_D
FONT_FC_TASK_E = Font(name="Microsoft YaHei", color="FF0070C0")

# 对齐
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
ALIGN_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_VCENTER = Alignment(vertical="center")
ALIGN_LEFT_VCENTER = Alignment(horizontal="left", vertical="center")
ALIGN_WRAP_VCENTER = Alignment(vertical="center", wrap_text=True)


def _apply_info_row(ws, row, col_a_val, col_d_val=None, col_e_val=None, col_f_val=None):
    """填写表头信息行(Row 2-6), 统一样式"""
    cell_a = ws.cell(row=row, column=1, value=col_a_val)
    cell_a.font = FONT_BASE
    cell_a.fill = FILL_INFO
    cell_a.alignment = ALIGN_VCENTER
    cell_a.border = THIN_BORDER

    # B, C 列也要有填充和边框
    for c in [2, 3]:
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_INFO
        cell.border = THIN_BORDER

    if col_d_val:
        cell_d = ws.cell(row=row, column=4, value=col_d_val)
        cell_d.font = FONT_INFO_LABEL
        cell_d.fill = FILL_INFO
        cell_d.alignment = ALIGN_CENTER
        cell_d.border = THIN_BORDER

    # E列
    cell_e = ws.cell(row=row, column=5)
    if col_e_val:
        cell_e.value = col_e_val
    cell_e.font = FONT_BASE
    cell_e.fill = FILL_INFO
    cell_e.alignment = ALIGN_VCENTER
    cell_e.border = THIN_BORDER

    if col_f_val:
        cell_f = ws.cell(row=row, column=6, value=col_f_val)
        cell_f.font = FONT_BASE
        cell_f.alignment = ALIGN_VCENTER


def _apply_task_row(ws, row, task_name, assignee, task_status, deadline, notes, row_fill):
    """填写任务行，根据row_fill选择不同字体组"""
    if row_fill == FILL_ROW_GREEN:
        fonts = (FONT_HW_TASK_A, FONT_HW_TASK_B, FONT_HW_TASK_C, FONT_HW_TASK_D, FONT_HW_TASK_E)
    elif row_fill == FILL_ROW_GRAY:
        fonts = (FONT_FC_TASK_A, FONT_FC_TASK_B, FONT_FC_TASK_C, FONT_FC_TASK_D, FONT_FC_TASK_E)
    else:
        fonts = (FONT_TASK_A, FONT_TASK_B, FONT_TASK_C, FONT_TASK_D, FONT_TASK_E)

    cols = [
        (1, task_name, fonts[0], ALIGN_CENTER),
        (2, assignee, fonts[1], ALIGN_CENTER),
        (3, task_status, fonts[2], ALIGN_CENTER),
        (4, deadline, fonts[3], ALIGN_CENTER),
        (5, notes, fonts[4], ALIGN_VCENTER),
    ]

    for col_idx, value, font, alignment in cols:
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = font
        cell.alignment = alignment
        cell.border = THIN_BORDER
        if row_fill:
            cell.fill = row_fill


def create_wbs(data: dict, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active

    # Sheet name
    version = data.get("version", "V1.0")
    date_str = data.get("date", "20260410")
    ws.title = f"WBS-{version}.{date_str}"

    # Column widths (精确匹配模板)
    ws.column_dimensions["A"].width = 93.71
    ws.column_dimensions["B"].width = 36.71
    ws.column_dimensions["C"].width = 19.71
    ws.column_dimensions["D"].width = 17.26
    ws.column_dimensions["E"].width = 132.46
    ws.column_dimensions["F"].width = 74.05

    row = 1
    project_name = data.get("project_name", "项目")
    group_name = data.get("group_name", "WBS模板")

    # === Row 1: 项目标题 ===
    title_text = f"{project_name}项目WBS-{version}.{date_str}({group_name})"
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    cell = ws.cell(row=row, column=1, value=title_text)
    cell.font = FONT_TITLE
    cell.fill = FILL_TITLE
    cell.alignment = ALIGN_CENTER
    ws.row_dimensions[row].height = 33
    row += 1

    # === Row 2: 部门 ===
    dept = data.get("department", "")
    _apply_info_row(ws, row,
        col_a_val=f"部       门：{dept}",
        col_d_val="禅道路径",
        col_f_val=data.get("board_type_ref", "Board_Type 命名规则对照表 V2.0"))
    row += 1

    # === Row 3: 项目成员 ===
    members = data.get("members", "")
    _apply_info_row(ws, row,
        col_a_val=f"项目成员：{members}",
        col_d_val="企业微盘",
        col_e_val=data.get("wepan_path", ""),
        col_f_val=data.get("spec_summary_ref", "产品规格汇总"))
    row += 1

    # === Row 4: 微信群组 ===
    wechat = data.get("wechat_group", "")
    _apply_info_row(ws, row,
        col_a_val=f"微信群组：{wechat}",
        col_d_val="规格文档",
        col_e_val=data.get("spec_doc", ""),
        col_f_val=data.get("burn_in_ref", ""))
    row += 1

    # === Row 5: 约束说明 ===
    constraints = data.get("constraints", "V1.0 开发阶段WBS，V2.0 基线WBS, V3.0 维护WBS ，WBS一周最少迭代一次")
    _apply_info_row(ws, row,
        col_a_val=f"约束说明：{constraints}",
        col_d_val="工作分支",
        col_e_val=data.get("branch", ""),
        col_f_val=data.get("ddr_flash_ref", "DDR&FLASH导入验证"))
    row += 1

    # === Row 6: 状态说明 ===
    project_status = data.get("project_status", "RUNNING")
    cell_a = ws.cell(row=row, column=1, value=f"状态说明：{project_status}")
    cell_a.font = FONT_BASE
    cell_a.fill = FILL_INFO
    cell_a.alignment = ALIGN_VCENTER

    for c in [2, 3]:
        cell = ws.cell(row=row, column=c)
        cell.fill = FILL_INFO

    cell_d = ws.cell(row=row, column=4, value="优先等级")
    cell_d.font = FONT_INFO_LABEL
    cell_d.fill = FILL_INFO
    cell_d.alignment = ALIGN_CENTER
    cell_d.border = THIN_BORDER

    cell_e = ws.cell(row=row, column=5, value=data.get("priority", "S级"))
    cell_e.font = FONT_PRIORITY
    cell_e.fill = FILL_INFO
    cell_e.alignment = ALIGN_LEFT_VCENTER
    row += 1

    # === Row 7: 硬件默认参数(F列) ===
    if data.get("hardware_notes"):
        cell_f = ws.cell(row=row, column=6, value=data["hardware_notes"])
        cell_f.font = FONT_HW_NOTES
        cell_f.alignment = ALIGN_WRAP_VCENTER
    ws.row_dimensions[row].height = 230.25
    row += 1

    # === Row 8: 列标题 ===
    headers = {
        1: ("任务(保证如下节点的执行纪律)", FONT_COL_HEADER, FILL_COL_HEADER),
        2: ("负责人", FONT_COL_HEADER, FILL_COL_HEADER),
        3: ("状态", FONT_COL_HEADER, FILL_COL_HEADER),
        4: ("预计结束时间", FONT_COL_HEADER, FILL_COL_HEADER_D),
        5: ("备注", FONT_COL_HEADER_E, FILL_COL_HEADER),
    }
    for col_idx, (text, font, fill) in headers.items():
        cell = ws.cell(row=row, column=col_idx, value=text)
        cell.font = font
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = ALIGN_CENTER
    ws.row_dimensions[row].height = 23.25
    row += 1

    # === 分区和任务 ===
    sections = data.get("sections", [])
    is_first_section = True

    for section in sections:
        section_name = section.get("name", "")
        tasks = section.get("tasks", [])

        # 分区标题行
        cell_a = ws.cell(row=row, column=1, value=section_name)
        cell_a.font = FONT_SECTION
        cell_a.fill = FILL_SECTION
        cell_a.alignment = ALIGN_CENTER

        # B-E列也设置分区样式
        for col in range(2, 6):
            c = ws.cell(row=row, column=col)
            c.fill = FILL_SECTION

        if is_first_section:
            # 第一个分区: 提示文字
            ws.cell(row=row, column=2, value='输入"@+人名"提及负责人').font = FONT_HINT
            ws.cell(row=row, column=2).fill = FILL_SECTION
            ws.cell(row=row, column=2).alignment = ALIGN_CENTER

            ws.cell(row=row, column=3, value="使用下拉列表显示").font = FONT_HINT
            ws.cell(row=row, column=3).fill = FILL_SECTION
            ws.cell(row=row, column=3).alignment = ALIGN_CENTER

            ws.cell(row=row, column=5, value="可输入内容或「插入」在线文档/图片到单元格中").font = FONT_HINT_E
            ws.cell(row=row, column=5).fill = FILL_SECTION
            ws.cell(row=row, column=5).alignment = ALIGN_CENTER

            is_first_section = False
        else:
            # 非第一分区: B/C也有分区标题字体, D列显示"预计结束时间"
            ws.cell(row=row, column=2).font = FONT_SECTION
            ws.cell(row=row, column=3).font = FONT_SECTION

            cell_d = ws.cell(row=row, column=4, value="预计结束时间")
            cell_d.font = FONT_SECTION_D
            cell_d.fill = FILL_SECTION
            cell_d.alignment = ALIGN_CENTER

        row += 1

        # 任务行
        for task in tasks:
            task_name = task.get("name", "")
            assignee = task.get("assignee", "")
            task_status = task.get("status", "待启动")
            deadline = task.get("deadline", "")
            notes = task.get("notes", "")

            # 根据任务名前缀确定行底色
            if task_name.startswith("【硬件】"):
                row_fill = FILL_ROW_GREEN
            elif task_name.startswith("【工厂】"):
                row_fill = FILL_ROW_GRAY
            else:
                row_fill = None

            _apply_task_row(ws, row, task_name, assignee, task_status, deadline, notes, row_fill)
            row += 1

        # 分区间空行 (保持样式一致: 空任务行样式)
        _apply_task_row(ws, row, "", "", "", "", "", None)
        row += 1

    # 状态列下拉验证
    dv = DataValidation(
        type="list",
        formula1='"待启动,进行中,已完成,取消"',
        allow_blank=True,
    )
    dv.error = "请选择有效状态"
    dv.errorTitle = "无效状态"
    dv.add(f"C9:C{row}")
    ws.add_data_validation(dv)

    # ============================================================
    # Sheet 2: 资源与性能
    # ============================================================
    rp = data.get("resource_performance", {})
    ws2 = wb.create_sheet(title="资源与性能")

    # 列宽 (精确匹配模板)
    ws2.column_dimensions["A"].width = 71.89
    ws2.column_dimensions["B"].width = 94.55
    ws2.column_dimensions["C"].width = 64.20

    r = 1
    # 标题行 (与WBS标题同样式)
    ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    cell = ws2.cell(row=r, column=1, value=f"{project_name}项目资源和性能")
    cell.font = FONT_TITLE
    cell.fill = FILL_TITLE
    cell.alignment = ALIGN_CENTER
    ws2.row_dimensions[r].height = 33
    r += 1

    # --- DDR 区块 ---
    # 区块标题行 (分区样式)
    ws2.cell(row=r, column=1, value="DDR").font = FONT_SECTION
    ws2.cell(row=r, column=1).fill = FILL_SECTION
    ws2.cell(row=r, column=2).fill = FILL_SECTION
    ws2.cell(row=r, column=3, value="可输入内容或「插入」在线文档/图片到单元格中").font = FONT_HINT_E
    ws2.cell(row=r, column=3).fill = FILL_SECTION
    r += 1

    ddr_items = rp.get("ddr_items", [
        "DDR满负载带宽",
        "DDR内存分布",
        "满负载内存media余量",
        "满负载内存OS余量",
    ])
    for item in ddr_items:
        ws2.cell(row=r, column=1, value=item).font = FONT_TASK_A
        ws2.cell(row=r, column=1).alignment = ALIGN_CENTER
        ws2.cell(row=r, column=1).border = THIN_BORDER
        ws2.cell(row=r, column=2).border = THIN_BORDER
        ws2.cell(row=r, column=3).border = THIN_BORDER
        r += 1

    # 空行 (带样式)
    for _ in range(2):
        ws2.cell(row=r, column=1).font = FONT_TASK_A
        ws2.cell(row=r, column=1).border = THIN_BORDER
        ws2.cell(row=r, column=2).border = THIN_BORDER
        ws2.cell(row=r, column=3).border = THIN_BORDER
        r += 1

    # --- NPU 使用率区块 ---
    npu_title = rp.get("npu_title", "NPU 使用率")
    npu_subtitle = rp.get("npu_subtitle", f"{project_name} AI性能测试")
    ws2.cell(row=r, column=1, value=npu_title).font = FONT_SECTION
    ws2.cell(row=r, column=1).fill = FILL_SECTION
    ws2.cell(row=r, column=2, value=npu_subtitle).font = FONT_SECTION_D
    ws2.cell(row=r, column=2).fill = FILL_SECTION
    ws2.cell(row=r, column=3).fill = FILL_SECTION
    r += 1

    npu_items = rp.get("npu_items", ["CNN频率"])
    for item in npu_items:
        ws2.cell(row=r, column=1, value=item).font = FONT_HW_TASK_A
        ws2.cell(row=r, column=1).alignment = ALIGN_CENTER
        ws2.cell(row=r, column=1).border = THIN_BORDER
        ws2.cell(row=r, column=2).border = THIN_BORDER
        ws2.cell(row=r, column=3).border = THIN_BORDER
        r += 1

    # AI 工作负载场景
    ai_scenario = rp.get("ai_scenario", "")
    if ai_scenario:
        ws2.cell(row=r, column=1, value=ai_scenario).font = FONT_HW_TASK_A
        ws2.cell(row=r, column=1).alignment = ALIGN_CENTER
        ws2.cell(row=r, column=1).border = THIN_BORDER
        ws2.cell(row=r, column=2).border = THIN_BORDER
        ws2.cell(row=r, column=3).border = THIN_BORDER
        r += 1

    # 空行
    for _ in range(2):
        ws2.cell(row=r, column=1).border = THIN_BORDER
        ws2.cell(row=r, column=2).border = THIN_BORDER
        ws2.cell(row=r, column=3).border = THIN_BORDER
        r += 1

    # --- FLASH 使用情况区块 ---
    ws2.cell(row=r, column=1, value="FLASH使用情况").font = FONT_SECTION
    ws2.cell(row=r, column=1).fill = FILL_SECTION
    ws2.cell(row=r, column=2).font = FONT_SECTION_D
    ws2.cell(row=r, column=2).fill = FILL_SECTION
    ws2.cell(row=r, column=3).fill = FILL_SECTION
    r += 1

    flash_items = rp.get("flash_items", [
        "rootfs 镜像格式、压缩方式",
        "rootfs分区余量",
        "app 镜像格式、压缩方式",
        "app分区余量",
        "downloade 分区大小和余量",
    ])
    for item in flash_items:
        ws2.cell(row=r, column=1, value=item).font = FONT_HW_TASK_A
        ws2.cell(row=r, column=1).alignment = ALIGN_CENTER
        ws2.cell(row=r, column=1).border = THIN_BORDER
        ws2.cell(row=r, column=2).border = THIN_BORDER
        ws2.cell(row=r, column=3).border = THIN_BORDER
        r += 1

    # 空行
    for _ in range(3):
        ws2.cell(row=r, column=1).border = THIN_BORDER
        ws2.cell(row=r, column=2).border = THIN_BORDER
        ws2.cell(row=r, column=3).border = THIN_BORDER
        r += 1

    # --- CPU 区块 ---
    ws2.cell(row=r, column=1, value="CPU").font = FONT_SECTION
    ws2.cell(row=r, column=1).fill = FILL_SECTION
    ws2.cell(row=r, column=2).font = FONT_SECTION_D
    ws2.cell(row=r, column=2).fill = FILL_SECTION
    ws2.cell(row=r, column=3).fill = FILL_SECTION
    r += 1

    cpu_items = rp.get("cpu_items", ["满负载CPU使用率"])
    for item in cpu_items:
        ws2.cell(row=r, column=1, value=item).font = FONT_HW_TASK_A
        ws2.cell(row=r, column=1).alignment = ALIGN_CENTER
        ws2.cell(row=r, column=1).border = THIN_BORDER
        ws2.cell(row=r, column=2).border = THIN_BORDER
        ws2.cell(row=r, column=3).border = THIN_BORDER
        r += 1

    wb.save(output_path)
    print(f"WBS已生成: {output_path}")


def main():
    if len(sys.argv) != 3:
        print(f"用法: {sys.argv[0]} <input.json> <output.xlsx>")
        sys.exit(1)

    input_path = sys.argv[1]
    output_path = sys.argv[2]

    with open(input_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    create_wbs(data, output_path)


if __name__ == "__main__":
    main()
